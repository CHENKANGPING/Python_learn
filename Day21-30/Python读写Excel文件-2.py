# pip install openpyxl
# 读取Excel
import datetime
import openpyxl
import openpyxl.workbook

wb = openpyxl.load_workbook('2022年股票数据.xlsx')

print(wb.sheetnames)

sheet = wb.worksheets[0]

print(sheet.dimensions)

print(sheet.max_row, sheet.max_column)

print(sheet.cell(3, 3).value)

print(sheet['C3'].value)

print(sheet['G255'].value)

print(sheet['A2:C5'])


for row_ch in range(2, sheet.max_row + 1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value: < 10d}', end='\t')
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(value, end='\t')
    print()
    
    
print('--------------------------------------------------')


# 写入Excel文件
import random
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = 'test scorces'

titles = ('name', 'Chinses', 'Math', 'English')
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)
    
names = ('ckp', 'hyf', 'zj', 'xxx')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))
        
wb.save('texts.xlsx')


print('--------------------------------------------------')


# import openpyxl
# from openpyxl.styles import Font, Alignment, Border, Side

# # 对其方式
# alignment = Alignment(horizontal='center', vertical='center')

# # 边框线条
# side = Side(color='ff7f50', style='mediumDashed')

# wb = openpyxl.load_workbook('texts.xlsx')
# sheet = wb.worksheets[0]

# # 调整行高和列宽
# sheet.row_dimensions[1].height = 30
# sheet.column_dimensions['E'].width = 120

# sheet['E1'] = '平均分'

# # 设置字体
# sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')

# # 设置对其方式
# sheet.cell(1,5).alignment = alignment

# # 设置单元格边框
# sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
# for i in range(2, 7):
#     # 公式计算平均分
#     sheet[f'E{i}'] = f'=average(B{i} : D{i})'
#     sheet.cell(i, 5).font = Font(size=12, color='4196e1', italic=True)
#     sheet.cell(i,5).alignment = alignment
    
# wb.save('tests-a.xlsx')


print('--------------------------------------------------')


# 生成统计图表
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook()  # 去掉 write_only=True
sheet = wb.active  # 使用默认的工作表，也可以创建新的工作表

rows = [
    ('类别', '销售A组', '销售B组'),
    ('手机', 40, 30),
    ('平板', 50, 60),
    ('笔记本', 80, 70),
    ('外围设备', 20, 10),
]

for row in rows:
    sheet.append(row)

chart = BarChart()
chart.type = 'col'
chart.style = 10
chart.title = '销售统计图'
chart.y_axis.title = '销量'
chart.x_axis.title = '商品类别'

# 修正 Reference 的范围
data = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=5)  # 数据范围
cats = Reference(sheet, min_col=1, min_row=2, max_row=5)  # 类别范围

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet.add_chart(chart, 'A10')  # 将图表插入到 A10 单元格

wb.save('demo.xlsx')  # 保存文件


